import os
import uuid
import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


def _rid(city="IND"):
    return "ESG-{}-{}-{}".format(city[:3].upper(), datetime.datetime.now().strftime('%Y'), uuid.uuid4().hex[:6].upper())


def _esc(text):
    """Escape & for ReportLab XML parsing (fixes H&M semicolon bug)."""
    if not text:
        return ""
    return str(text).replace("&", "&amp;")


# Color constants
PC = colors.HexColor('#2D6A4F')   # Pass/Good
EC = colors.HexColor('#10B981')   # Excellent
WC = colors.HexColor('#B8860B')   # Warning/Moderate
FC = colors.HexColor('#9B2226')   # Fail/Poor
CC = colors.HexColor('#7F1D1D')   # Critical


def _status_color(status):
    """Get color for 5-level fuzzy status."""
    s = str(status).lower()
    if s == "excellent":
        return EC
    elif s == "good":
        return PC
    elif s == "moderate":
        return WC
    elif s == "poor":
        return FC
    elif s == "critical":
        return CC
    return colors.black


def generate_pdf_report(company_data, kpis, verification_checks, gap_analysis, suggestions_data, fuzzy_data=None):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm,
                            rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    ts = ParagraphStyle('T', parent=styles['Title'], fontSize=20, spaceAfter=4,
                        textColor=colors.HexColor('#1B4332'), fontName='Helvetica-Bold')
    hs = ParagraphStyle('H', parent=styles['Heading2'], fontSize=12, spaceAfter=6,
                        spaceBefore=14, textColor=colors.HexColor('#1B4332'), fontName='Helvetica-Bold')
    bs = ParagraphStyle(
        'B', parent=styles['Normal'], fontSize=9, spaceAfter=3, leading=12)
    ss = ParagraphStyle(
        'S', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#666'), leading=10)
    els = []

    city = company_data.get("city", company_data.get("state", "IND"))
    rid = _rid(city)
    now = datetime.datetime.now().strftime("%d %B %Y, %I:%M %p")
    buyer = _esc(company_data.get("buyer_name", "N/A"))

    # ── Gather all scores up front ──
    from components.calculator import get_overall_score
    sc = get_overall_score(kpis)
    overall_data = None
    if fuzzy_data and fuzzy_data.get("overall"):
        overall_data = fuzzy_data["overall"]
        sc = overall_data.get("score", sc)

    sc_col = '#10B981' if sc >= 85 else '#2D6A4F' if sc >= 70 else '#B8860B' if sc >= 40 else '#9B2226' if sc >= 20 else '#7F1D1D'
    sc_lab = 'Excellent' if sc >= 85 else 'Good Standing' if sc >= 70 else 'Needs Improvement' if sc >= 40 else 'Poor' if sc >= 20 else 'Critical'

    gw_data = (fuzzy_data or {}).get("greenwash", {})
    gw_score = gw_data.get("risk_score", 0)
    gw_label = gw_data.get("risk_label", "N/A")
    gw_col = '#9B2226' if gw_label == "High" else '#B8860B' if gw_label == "Medium" else '#2D6A4F'

    fc_data = (fuzzy_data or {}).get("confidence", {})
    conf_score = fc_data.get("score", fc_data.get("overall_score", 0))
    conf_label = fc_data.get("label", fc_data.get("overall_label", "N/A"))
    conf_col = '#2D6A4F' if conf_score >= 75 else '#B8860B' if conf_score >= 50 else '#9B2226'

    readiness_score = (gap_analysis or {}).get("readiness_score", 0)
    readiness_col = '#2D6A4F' if readiness_score >= 75 else '#B8860B' if readiness_score >= 50 else '#9B2226'
    readiness_lab = 'Ready' if readiness_score >= 75 else 'Partial' if readiness_score >= 50 else 'Not Ready'

    # ══════════════════════════════════════════
    # PAGE 1 — COVER + FOUR SCORES
    # ══════════════════════════════════════════

    els.append(Paragraph("ESGVerify", ts))
    els.append(Paragraph("BRSR Core ESG Assessment Report", bs))
    els.append(Spacer(1, 3*mm))
    meta = [
        ["Report ID:", rid, "Date:", now],
        ["Company:", _esc(company_data.get("company_name", "N/A")),
         "GSTIN:", company_data.get("gstin", "N/A")],
        ["Industry:", _esc(company_data.get("industry", "N/A")),
         "State:", company_data.get("state", "N/A")],
        ["Workers:", str(company_data.get("workers", "N/A")),
         "Target Buyer:", buyer],
    ]
    mt = Table(meta, colWidths=[55, 150, 55, 150])
    mt.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#666')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    els.append(mt)
    els.append(Spacer(1, 4*mm))
    els.append(HRFlowable(width="100%", thickness=0.5,
               color=colors.HexColor('#CCC')))
    els.append(Spacer(1, 4*mm))

    # 2×2 score grid
    def _score_cell(title, score_val, col_hex, label_str):
        return [
            Paragraph(
                '<font size="8" color="#555"><b>{}</b></font>'.format(title), bs),
            Paragraph('<font size="20" color="{}"><b>{}</b></font><font size="9" color="#888"> /100</font>'.format(
                col_hex, int(round(score_val))), bs),
            Paragraph(
                '<font size="8" color="{}"><b>{}</b></font>'.format(col_hex, label_str), bs),
        ]

    grid_bg_esg = '#F0FFF4' if sc >= 70 else '#FFFBEB' if sc >= 40 else '#FFF5F5'
    grid_bg_gw = '#FFF5F5' if gw_label == "High" else '#FFFBEB' if gw_label == "Medium" else '#F0FFF4'
    grid_bg_conf = '#F0FFF4' if conf_score >= 75 else '#FFFBEB' if conf_score >= 50 else '#FFF5F5'
    grid_bg_ready = '#F0FFF4' if readiness_score >= 75 else '#FFFBEB' if readiness_score >= 50 else '#FFF5F5'

    grid_data = [
        [_score_cell("ESG SCORE", sc, sc_col, sc_lab),
         _score_cell("GREENWASH RISK", gw_score, gw_col, gw_label)],
        [_score_cell("DATA CONFIDENCE", conf_score, conf_col, conf_label),
         _score_cell("BUYER READINESS", readiness_score, readiness_col, readiness_lab)],
    ]
    grid_table = Table(grid_data, colWidths=[240, 240])
    grid_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#DDD')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDD')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor(grid_bg_esg)),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor(grid_bg_gw)),
        ('BACKGROUND', (0, 1), (0, 1), colors.HexColor(grid_bg_conf)),
        ('BACKGROUND', (1, 1), (1, 1), colors.HexColor(grid_bg_ready)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    els.append(Paragraph("ASSESSMENT SCORES", hs))
    els.append(grid_table)
    els.append(Spacer(1, 6*mm))

    # Assessment Summary — plain English, all 4 scores
    esg_sent = "This company achieved an ESG score of {}/100 ({}), {}.".format(
        int(round(sc)), sc_lab,
        "reflecting strong ESG practices across all pillars" if sc >= 70
        else "indicating performance that needs improvement in key areas" if sc >= 40
        else "highlighting significant gaps that require immediate attention")
    gw_sent = "Greenwash risk is rated {}/100 ({}), {}.".format(
        int(round(gw_score)), gw_label,
        "indicating a high likelihood of unsubstantiated claims that require attention" if gw_label == "High"
        else "suggesting some claims may need stronger supporting evidence" if gw_label == "Medium"
        else "indicating the company's ESG claims appear well-supported")
    conf_sent = "Data confidence stands at {}/100 ({}), {}.".format(
        int(round(conf_score)), conf_label,
        "with most inputs verified and reliable" if conf_score >= 75
        else "suggesting partial verification with some data gaps" if conf_score >= 50
        else "revealing significant data quality concerns")
    ready_sent = "Buyer readiness for {} is {}/100 ({}), {}.".format(
        buyer, int(round(readiness_score)), readiness_lab,
        "meaning the company fully meets buyer requirements" if readiness_score >= 75
        else "meaning core requirements are met but specific gaps remain" if readiness_score >= 50
        else "meaning significant gaps exist before meeting buyer standards")

    els.append(Paragraph("ASSESSMENT SUMMARY", hs))
    els.append(Paragraph(
        _esc(" ".join([esg_sent, gw_sent, conf_sent, ready_sent])),
        ParagraphStyle('AS', parent=bs, fontSize=9, leading=14)))

    # ══════════════════════════════════════════
    # PAGE 2 — ESG SCORE DETAIL
    # ══════════════════════════════════════════
    els.append(PageBreak())
    els.append(Paragraph("ESG SCORE DETAIL", hs))
    els.append(Paragraph(
        '<font size="28" color="{}"><b>{}</b></font>'
        '<font size="12" color="#666"> / 100 — {}</font>'.format(sc_col, int(round(sc)), sc_lab), bs))
    els.append(Spacer(1, 4*mm))

    # Pillar breakdown
    if overall_data:
        pillar_scores = overall_data.get("pillar_scores", {})
        weights = overall_data.get("weights_used", {})
        if pillar_scores:
            e_sc = int(round(pillar_scores.get("Environmental", 0)))
            s_sc = int(round(pillar_scores.get("Social", 0)))
            g_sc = int(round(pillar_scores.get("Governance", 0)))
            e_wt = int((weights or {}).get("Environmental", 0.55) * 100)
            s_wt = int((weights or {}).get("Social", 0.30) * 100)
            g_wt = int((weights or {}).get("Governance", 0.15) * 100)
            e_col = '#10B981' if e_sc >= 85 else '#2D6A4F' if e_sc >= 70 else '#B8860B' if e_sc >= 40 else '#9B2226'
            s_col = '#10B981' if s_sc >= 85 else '#2D6A4F' if s_sc >= 70 else '#B8860B' if s_sc >= 40 else '#9B2226'
            g_col = '#10B981' if g_sc >= 85 else '#2D6A4F' if g_sc >= 70 else '#B8860B' if g_sc >= 40 else '#9B2226'
            pillar_rows = [
                ["Pillar", "Weight", "Score"],
                [Paragraph('<font color="{}"><b>Environmental</b></font>'.format(e_col), bs),
                 "{}%".format(e_wt), "{}/100".format(e_sc)],
                [Paragraph('<font color="{}"><b>Social</b></font>'.format(s_col), bs),
                 "{}%".format(s_wt), "{}/100".format(s_sc)],
                [Paragraph('<font color="{}"><b>Governance</b></font>'.format(g_col), bs),
                 "{}%".format(g_wt), "{}/100".format(g_sc)],
            ]
            els.append(Paragraph("PILLAR BREAKDOWN", hs))
            pt = Table(pillar_rows, colWidths=[220, 70, 70])
            pt.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4332')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDD')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#F5F5F0')]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ]))
            els.append(pt)
            els.append(Spacer(1, 4*mm))

    # KPI results table
    els.append(Paragraph("BRSR CORE KPI RESULTS", hs))
    rows = [["No.", "KPI", "Value", "Unit", "Score", "Status", "Data"]]
    for k in kpis:
        fuzzy = k.get("fuzzy", {})
        dominant = fuzzy.get("dominant", k.get("status", "moderate")).title()
        rows.append([
            str(k["kpi_number"]),
            _esc(k["kpi_name"]),
            k["display_value"],
            _esc(k["unit"][:28]),
            "{}/100".format(k['score']),
            dominant,
            k.get("confidence", "measured").title(),
        ])
    kt = Table(rows, colWidths=[18, 90, 50, 100, 35, 48, 48])
    sty = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4332')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#F5F5F0')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDD')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
    ]
    for i, k in enumerate(kpis, 1):
        fuzzy = k.get("fuzzy", {})
        dominant = fuzzy.get("dominant", k.get("status", "moderate"))
        c = _status_color(dominant)
        sty.append(('TEXTCOLOR', (5, i), (5, i), c))
        sty.append(('FONTNAME', (5, i), (5, i), 'Helvetica-Bold'))
    kt.setStyle(TableStyle(sty))
    els.append(kt)
    els.append(Spacer(1, 5*mm))

    # Areas to improve — KPIs below 70
    weak_kpis = [k for k in kpis if k["score"] < 70]
    if weak_kpis:
        els.append(Paragraph("AREAS TO IMPROVE", hs))
        for k in weak_kpis:
            fuzzy = k.get("fuzzy", {})
            dominant = fuzzy.get("dominant", k.get("status", "moderate"))
            c_hex = ('#10B981' if dominant == "excellent" else '#2D6A4F' if dominant == "good"
                     else '#B8860B' if dominant == "moderate" else '#9B2226' if dominant == "poor" else '#7F1D1D')
            els.append(Paragraph(
                '<font color="{}">- {}/100 &mdash; <b>{}</b></font>'.format(
                    c_hex, k["score"], _esc(k["kpi_name"])), ss))
        els.append(Spacer(1, 3*mm))

    # ══════════════════════════════════════════
    # PAGE 3 — GREENWASH & CONFIDENCE
    # ══════════════════════════════════════════
    els.append(PageBreak())

    # Greenwash Risk
    els.append(Paragraph("GREENWASH RISK", hs))
    els.append(Paragraph(
        '<font size="24" color="{}"><b>{}</b></font>'
        '<font size="11" color="#666"> / 100 &mdash; {}</font>'.format(
            gw_col, int(round(gw_score)), gw_label), bs))
    els.append(Spacer(1, 3*mm))

    explanations = gw_data.get("explanations", [])
    if explanations:
        els.append(Paragraph('<b>Why this risk level:</b>',
                             ParagraphStyle('GWH', parent=bs, fontSize=9)))
        for exp in explanations[:5]:
            els.append(Paragraph('- {}'.format(_esc(exp)), ss))
        els.append(Spacer(1, 2*mm))

    signals = gw_data.get("signals", gw_data.get("active_signals", {}))
    if signals:
        els.append(Paragraph(
            '<font size="8" color="#444">Signal degrees (0 = no risk, 1 = maximum):</font>', bs))
        for sig_name, sig_degree in signals.items():
            if sig_degree > 0.01:
                sig_col = '#9B2226' if sig_degree > 0.5 else '#B8860B' if sig_degree > 0.2 else '#2D6A4F'
                els.append(Paragraph(
                    '<font size="7" color="{}">  {} : {:.2f}</font>'.format(
                        sig_col, sig_name.replace("_", " ").title(), sig_degree), ss))
        els.append(Spacer(1, 4*mm))

    # Data Confidence
    els.append(HRFlowable(width="100%", thickness=0.5,
               color=colors.HexColor('#CCC')))
    els.append(Paragraph("DATA CONFIDENCE", hs))
    els.append(Paragraph(
        '<font size="24" color="{}"><b>{}</b></font>'
        '<font size="11" color="#666"> / 100 &mdash; {}</font>'.format(
            conf_col, int(round(conf_score)), conf_label), bs))
    els.append(Spacer(1, 2*mm))

    if verification_checks:
        passed = sum(1 for ch in verification_checks if ch["status"] == "pass")
        warned = sum(1 for ch in verification_checks if ch["status"] == "warn")
        failed = sum(1 for ch in verification_checks if ch["status"] == "fail")
        els.append(Paragraph(
            '<font color="#2D6A4F"><b>{} passed</b></font>, '
            '<font color="#B8860B"><b>{} warnings</b></font>, '
            '<font color="#9B2226"><b>{} failed</b></font>'.format(passed, warned, failed), bs))
        els.append(Spacer(1, 2*mm))

    conf_detail = fc_data.get("detail", fc_data.get("summary", ""))
    if conf_detail:
        els.append(Paragraph(_esc(str(conf_detail)), ss))
    els.append(Spacer(1, 3*mm))

    # Show all verification checks
    if verification_checks:
        els.append(Paragraph("VERIFICATION CHECKS", hs))
        for ch in verification_checks:
            c = PC if ch["status"] == "pass" else WC if ch["status"] == "warn" else FC
            icon = "PASS" if ch["status"] == "pass" else "WARN" if ch["status"] == "warn" else "FAIL"
            bg = '#E8F5E9' if ch["status"] == "pass" else '#FFF8E1' if ch["status"] == "warn" else '#FFEBEE'
            vtbl = Table(
                [[icon, '{} — {}'.format(_esc(ch["title"]), _esc(ch["message"]))]], colWidths=[35, 354])
            vtbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(bg)),
                ('TEXTCOLOR', (0, 0), (0, 0), c),
                ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#374151')),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('BOX', (0, 0), (-1, -1), 0.3, colors.HexColor('#DDD')),
            ]))
            els.append(vtbl)
            els.append(Spacer(1, 1*mm))
    els.append(Spacer(1, 4*mm))

    # ══════════════════════════════════════════
    # PAGE 4 — BUYER READINESS
    # ══════════════════════════════════════════
    els.append(PageBreak())
    els.append(Paragraph("BUYER READINESS &mdash; {}".format(buyer), hs))
    els.append(Paragraph(
        '<font size="24" color="{}"><b>{}</b></font>'
        '<font size="11" color="#666"> / 100 &mdash; {}</font>'.format(
            readiness_col, int(round(readiness_score)), readiness_lab), bs))
    els.append(Spacer(1, 3*mm))

    if gap_analysis:
        summary_text = gap_analysis.get("summary", "")
        if summary_text:
            els.append(Paragraph(_esc(summary_text), bs))
            els.append(Spacer(1, 3*mm))

        if gap_analysis.get("gaps"):
            els.append(Paragraph("GAP ANALYSIS", hs))
            for g in gap_analysis["gaps"]:
                g_status = g["status"].upper()
                g_col = PC if g["status"] == "met" else WC if g["status"] == "borderline" else FC
                g_bg = '#E8F5E9' if g["status"] == "met" else '#FFF8E1' if g["status"] == "borderline" else '#FFEBEE'

                # KPI name + status badge
                els.append(Paragraph(
                    '<font color="{}"><b>{}</b></font>'.format(
                        g_col.hexval() if hasattr(g_col, 'hexval') else '#B8860B',
                        _esc(g["kpi_name"])),
                    ParagraphStyle('GN', parent=bs, fontSize=10, spaceBefore=6)))

                # Your value vs buyer requirement — side by side
                gap_row = Table(
                    [["Your Value", "Buyer Requires", "Status"],
                     [Paragraph(_esc(g["current_value"]), ss),
                      Paragraph(_esc(g.get("buyer_requirement", "N/A")), ss),
                      g_status]],
                    colWidths=[150, 180, 59])
                gap_row.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5F5F0')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('TEXTCOLOR', (2, 1), (2, 1), g_col),
                    ('FONTNAME', (2, 1), (2, 1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#DDD')),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                els.append(gap_row)

                # Gap detail — full width paragraph
                if g.get("gap_detail"):
                    els.append(Paragraph(
                        _esc(g["gap_detail"]),
                        ParagraphStyle('GD', parent=ss, fontSize=7, leading=10,
                                       textColor=colors.HexColor('#4B5563'),
                                       spaceBefore=2, spaceAfter=4)))
                els.append(Spacer(1, 2*mm))
            els.append(Spacer(1, 3*mm))
            gsty = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D3748')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#F5F5F0')]),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDD')),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]
            for i, g in enumerate(gap_analysis["gaps"], 1):
                c = PC if g["status"] == "met" else WC if g["status"] == "borderline" else FC
                gsty.append(('TEXTCOLOR', (3, i), (3, i), c))
                gsty.append(('FONTNAME', (3, i), (3, i), 'Helvetica-Bold'))

        priorities = gap_analysis.get("buyer_priorities", [])
        if priorities:
            els.append(Paragraph("BUYER PRIORITIES", hs))
            for p in priorities:
                els.append(Paragraph('- {}'.format(_esc(str(p))), ss))
            els.append(Spacer(1, 3*mm))

    # ─── DISCLAIMER ───
    els.append(Spacer(1, 5*mm))
    els.append(HRFlowable(width="100%", thickness=0.5,
               color=colors.HexColor('#CCC')))
    els.append(
        Paragraph("<b>Report ID:</b> {} | <b>Generated:</b> {}".format(rid, now), ss))
    els.append(Paragraph(
        "<b>Disclaimer:</b> Generated from self-declared data. Verification uses GSTIN checksum, "
        "industry benchmarks, cross-field consistency checks, and fuzzy logic confidence scoring. "
        "ESG pillar weights are dynamically derived from buyer's published ESG requirements. "
        "Final verification by buyer/auditor per SEBI BRSR guidelines.",
        ParagraphStyle('D', parent=ss, textColor=colors.HexColor('#999'), fontSize=7)))

    doc.build(els)
    buf.seek(0)
    return buf, rid


def generate_improvement_pdf(company_data, kpis, suggestions_data, gap_analysis):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm,
                            rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    ts = ParagraphStyle('T', parent=styles['Title'], fontSize=18, spaceAfter=4,
                        textColor=colors.HexColor('#1B4332'), fontName='Helvetica-Bold')
    hs = ParagraphStyle('H', parent=styles['Heading2'], fontSize=12, spaceAfter=6,
                        spaceBefore=10, textColor=colors.HexColor('#1B4332'), fontName='Helvetica-Bold')
    bs = ParagraphStyle(
        'B', parent=styles['Normal'], fontSize=9, spaceAfter=3, leading=12)
    ss = ParagraphStyle(
        'S', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#444'), leading=11)
    els = []
    now = datetime.datetime.now().strftime("%d %B %Y")
    els.append(Paragraph("ESGVerify - Improvement Roadmap", ts))
    els.append(Paragraph("{} | {}".format(
        _esc(company_data.get('company_name', 'Company')), now), bs))
    els.append(Spacer(1, 4*mm))
    els.append(HRFlowable(width="100%", thickness=0.5,
               color=colors.HexColor('#CCC')))

    # Weak areas
    els.append(Paragraph("AREAS REQUIRING IMPROVEMENT", hs))
    weak = [k for k in kpis if k["score"] < 70]
    if weak:
        for k in weak:
            fuzzy = k.get("fuzzy", {})
            dominant = fuzzy.get("dominant", k.get("status", "moderate"))
            c_hex = '#10B981' if dominant == "excellent" else '#2D6A4F' if dominant == "good" else '#B8860B' if dominant == "moderate" else '#9B2226' if dominant == "poor" else '#7F1D1D'
            els.append(Paragraph(
                '<font color="{}"><b>{}</b></font> — Score: {}/100 — {} {} — Classification: {}'.format(
                    c_hex, _esc(k["kpi_name"]), k["score"], k["display_value"], _esc(k["unit"]), dominant.title()), bs))
    else:
        els.append(Paragraph(
            "All KPIs are in good standing. Focus on maintaining performance.", bs))

    # Suggestions
    if suggestions_data and suggestions_data.get("suggestions"):
        els.append(Paragraph("PRIORITIZED ACTIONS", hs))
        for s in suggestions_data["suggestions"]:
            els.append(Paragraph(
                '<b>Priority {}: {}</b>'.format(s.get("priority", ""), _esc(s["title"])), bs))
            els.append(Paragraph(_esc(s["description"]), ss))
            els.append(Paragraph(
                'Cost: {} | Saving: {} | Payback: {}'.format(
                    _esc(s.get("estimated_cost", "N/A")),
                    _esc(s.get("estimated_annual_saving", "N/A")),
                    _esc(s.get("payback_period", "N/A"))), ss))
            els.append(Spacer(1, 2*mm))

    # Roadmap
    if suggestions_data and suggestions_data.get("roadmap"):
        els.append(Paragraph("12-WEEK ACTION PLAN", hs))
        for period, actions in suggestions_data["roadmap"].items():
            label = period.replace("_", " ").replace("week", "Week").title()
            els.append(Paragraph('<b>{}</b>'.format(label), bs))
            for a in actions:
                els.append(Paragraph('  - {}'.format(_esc(a)), ss))
            els.append(Spacer(1, 1*mm))

    doc.build(els)
    buf.seek(0)
    return buf


def generate_excel_report(company_data, kpis, verification_checks, gap_analysis, suggestions_data):
    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="1B4332",
                        end_color="1B4332", fill_type="solid")
    gf = PatternFill(start_color="D8F3DC",
                     end_color="D8F3DC", fill_type="solid")
    wf = PatternFill(start_color="FFF3CD",
                     end_color="FFF3CD", fill_type="solid")
    rf = PatternFill(start_color="FFCCD5",
                     end_color="FFCCD5", fill_type="solid")

    ws1 = wb.active
    ws1.title = "Summary"
    for r in [
        ["ESGVerify - BRSR Core Report"],
        ["Date", datetime.datetime.now().strftime("%d %B %Y")],
        [],
        ["COMPANY"],
        ["Name", company_data.get("company_name", "")],
        ["GSTIN", company_data.get("gstin", "")],
        ["Industry", company_data.get("industry", "")],
        ["State", company_data.get("state", "")],
        ["Workers", company_data.get("workers", "")],
        ["Buyer", company_data.get("buyer_name", "")],
    ]:
        ws1.append(r)

    ws2 = wb.create_sheet("KPIs")
    ws2.append(["No", "KPI", "Ref", "Value", "Unit", "Score",
               "Status", "Fuzzy Class", "Confidence", "Formula", "Source"])
    for c in ws2[1]:
        c.font = hf
        c.fill = hfill
    for k in kpis:
        fuzzy = k.get("fuzzy", {})
        dominant = fuzzy.get("dominant", k.get("status", "moderate")).title()
        ws2.append([
            k["kpi_number"], k["kpi_name"], k["brsr_ref"], k["display_value"],
            k["unit"], k["score"], k["status"].title(), dominant,
            k.get("confidence", "").title(), k["formula"], k["formula_source"],
        ])
        rn = ws2.max_row
        ws2.cell(
            row=rn, column=7).fill = gf if k["status"] == "good" else wf if k["status"] == "moderate" else rf

    if verification_checks:
        ws3 = wb.create_sheet("Verification")
        ws3.append(["Status", "Check", "Details", "Category"])
        for c in ws3[1]:
            c.font = hf
            c.fill = hfill
        for ch in verification_checks:
            ws3.append([ch["status"].upper(), ch["title"],
                       ch["message"], ch.get("category", "")])
            rn = ws3.max_row
            ws3.cell(
                row=rn, column=1).fill = gf if ch["status"] == "pass" else wf if ch["status"] == "warn" else rf

    if gap_analysis and gap_analysis.get("gaps"):
        ws4 = wb.create_sheet("Gap Analysis")
        ws4.append(["KPI", "Current", "Required",
                   "Status", "Detail", "Priority"])
        for c in ws4[1]:
            c.font = hf
            c.fill = hfill
        for g in gap_analysis["gaps"]:
            ws4.append([
                g["kpi_name"], g["current_value"],
                g.get("buyer_requirement", ""), g["status"].upper(),
                g["gap_detail"], g.get("priority", ""),
            ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
